#!/usr/bin/env python3
from __future__ import annotations
from pathlib import Path
from typing import List, Optional
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# ========= CONFIG =========

WORKBOOK_PATH = r"C:\Users\corri\Kinsmen Group\P25025 Plains Adept to Meridian Migration Project - Documents\6. Data Migration\PROD Location Loadsheets\Masters\Data Issues\Batch 21\Batch 21 Fixes.xlsx"

INPUT_SHEET = "Documents"
OUTPUT_SHEET = "Rendition_Actions"
TARGET_EXTS = ["dwg", "doc", "docx", "xls", "xlsm", "xlsx", "gp4"]
REVIEW_TAG = "REVIEW: MATCH UNCLEAR"
HEADER_RENAME_MAP = {
    "stack_id": "Stack ID",
    "document number": "Document Number",
    "temp revision number": "Temp Revision Number",
    "legacy version number": "Legacy Version Number",
    "ext": "Ext",
    "source_path": "Source Path",
    "rendition_path": "Rendition Path",
    "islatest": "isLatest",
}

# ========= HELPERS =========
def _norm_rev(val) -> str:
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return ""
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else s
    except ValueError:
        pass
    return str(int(s)) if s.isdigit() else s.upper()


def _truthy(x) -> bool:
    if pd.isna(x):
        return False
    if isinstance(x, bool):
        return x
    return str(x).strip().lower() in {"true", "yes", "y", "1", "t"}


def _is_adlib_path(s: pd.Series) -> pd.Series:
    return s.astype(str).str.contains(r"\\ADLib_", case=False, regex=True).fillna(False)


def _pick_pdf_row_id(gdf: pd.DataFrame) -> Optional[int]:
    if gdf.empty:
        return None
    g = gdf.copy()
    g["_is_adlib"] = _is_adlib_path(g["Source Path"]).astype(int)
    g = g.sort_values(["Legacy Version Number", "_is_adlib", "row_id"])
    return int(g.iloc[-1]["row_id"])


def _pick_target_row_id(gdf: pd.DataFrame, target_exts: List[str]) -> Optional[int]:
    if gdf.empty:
        return None
    pr = {e: i for i, e in enumerate([e.lower() for e in target_exts])}
    g = gdf.copy()
    g["ext_rank"] = g["Ext"].str.lower().map(pr).fillna(len(pr)).astype(int)
    g = g.sort_values(["ext_rank", "Legacy Version Number", "row_id"])
    return int(g.iloc[0]["row_id"])


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns={c: c.strip() for c in df.columns})
    df = df.rename(columns={c: HEADER_RENAME_MAP[c.lower()] for c in df.columns if c.lower() in HEADER_RENAME_MAP})

    required = {"Stack ID", "Document Number", "Temp Revision Number", "Legacy Version Number", "Ext", "Source Path", "Rendition Path"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns: {sorted(missing)}")

    df = df.copy()
    for c in ["Ext", "Temp Revision Number", "Document Number", "Stack ID", "Source Path", "Rendition Path"]:
        df[c] = df[c].astype(object).where(pd.notna(df[c]), "")
        df[c] = df[c].astype(str).str.strip()

    df["isLatest"] = df["isLatest"].apply(_truthy) if "isLatest" in df.columns else False
    df["Temp Revision Number"] = df["Temp Revision Number"].apply(_norm_rev)
    df["Ext"] = df["Ext"].str.lower().str.replace(".", "", regex=False)
    df["Legacy Version Number"] = pd.to_numeric(df["Legacy Version Number"], errors="coerce").fillna(-1).astype(int)
    df["row_id"] = df.index
    return df


def tidy(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["Rendition Path", "Source Path", "Stack ID", "Document Number", "Temp Revision Number"]:
        df[col] = df[col].replace({"nan": "", "NaN": ""})
    return df

# ========= WRITERS =========
def _write_update_with_path(out: pd.DataFrame, rid: int, action: str, copied_path: str):
    out.at[rid, "Action"] = action
    out.at[rid, "Reason"] = ""
    out.at[rid, "Rendition Path"] = copied_path or ""


def _write_reason_only(out: pd.DataFrame, rid: int, action: str, reason: str):
    out.at[rid, "Action"] = action
    out.at[rid, "Reason"] = reason
    out.at[rid, "Rendition Path"] = reason


# ========= PAIRING HELPERS =========
def _attempt_pairing(out: pd.DataFrame, pdfs: pd.DataFrame, tgts: pd.DataFrame, pair_action: str) -> None:
    if len(pdfs) and len(tgts):
        pdf_id = _pick_pdf_row_id(pdfs)
        tgt_id = _pick_target_row_id(tgts, TARGET_EXTS)

        if pdf_id is not None and tgt_id is not None:
            src = out.at[pdf_id, "Source Path"]
            _write_update_with_path(out, tgt_id, pair_action, src)
            _write_reason_only(out, pdf_id, "REMOVE", f"REMOVE: PDF consumed by {pair_action.split('(')[0].strip()} target")
            out.at[pdf_id, "_paired"] = True
            out.at[tgt_id, "_paired"] = True

        for idx in pdfs.index.difference([pdf_id]):
            _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra PDF in LVN pairing group")
        for idx in tgts.index.difference([tgt_id]):
            _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra target in LVN pairing group")
    else:
        for rid in pdfs["row_id"].tolist() + tgts["row_id"].tolist():
            _write_reason_only(out, int(rid), "REVIEW", f"{REVIEW_TAG}: LVN pairing picker failure")


def _handle_lone_pdfs_and_unmatched_targets(out: pd.DataFrame, grev: pd.DataFrame, target_exts: List[str]) -> None:
    g_unpaired = out.loc[grev.index]
    g_unpaired = g_unpaired[~g_unpaired["_paired"].fillna(False)]

    pdfs = g_unpaired[g_unpaired["Ext"].str.lower().eq("pdf")]
    tgts = g_unpaired[g_unpaired["Ext"].isin(target_exts)]

    if len(pdfs):
        pdf_id = _pick_pdf_row_id(pdfs)
        if pdf_id is not None:
            _write_update_with_path(out, pdf_id, "UPDATE RENDITION (LONE PDF)", out.at[pdf_id, "Source Path"])
            out.at[pdf_id, "_paired"] = True
        for idx in pdfs.index.difference([pdf_id] if pdf_id is not None else []):
            _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra PDF without eligible target (LONE PDF processing)")

    if len(tgts):
        for rid in tgts["row_id"]:
            if not out.at[int(rid), "_paired"]:
                is_latest_status = "TRUE" if out.at[int(rid), "isLatest"] else "FALSE"
                _write_reason_only(out, int(rid), "REVIEW", f"{REVIEW_TAG}: Unmatched Target ({is_latest_status})")
                out.at[int(rid), "_paired"] = True


# ========= CORE MATCHING =========
def _prepass_explicit_true_pair(out: pd.DataFrame, grev: pd.DataFrame, target_exts: List[str]) -> None:
    latest = grev[grev["isLatest"]]
    pdfs = latest[latest["Ext"].eq("pdf")]
    tgts = latest[latest["Ext"].isin(target_exts)]

    if len(pdfs) == 1 and len(tgts) == 1:
        pdf_id = int(pdfs.iloc[0]["row_id"])
        tgt_id = int(tgts.iloc[0]["row_id"])
        src = out.at[pdf_id, "Source Path"]

        _write_update_with_path(out, tgt_id, "UPDATE RENDITION (Explicit True Pair)", src)
        _write_reason_only(out, pdf_id, "REMOVE", "REMOVE: PDF consumed by Explicit True Pair target")
        out.at[pdf_id, "_paired"] = True
        out.at[tgt_id, "_paired"] = True


def _prepass_strict_lvn_pairing(out: pd.DataFrame, grev: pd.DataFrame, is_latest_val: bool) -> None:
    g_subset = grev[grev["isLatest"].eq(is_latest_val)]
    g_unpaired = g_subset[~g_subset["_paired"].fillna(False)]

    for lvn, g_lvn in g_unpaired.groupby("Legacy Version Number", sort=False):
        if pd.isna(lvn):
            continue
        pdfs = g_lvn[g_lvn["Ext"].eq("pdf")]
        tgts = g_lvn[g_lvn["Ext"].isin(TARGET_EXTS)]

        if len(pdfs) and len(tgts):
            action = f"UPDATE RENDITION (LVN match, isLatest={is_latest_val})"
            _attempt_pairing(out, pdfs, tgts, action)


def apply_spreadsheet_updates(df: pd.DataFrame, target_exts: List[str]) -> pd.DataFrame:
    out = df.copy()
    out["Action"] = "NO CHANGE"
    out["Reason"] = ""
    out["_paired"] = False

    for _, grev in out.groupby(["Stack ID", "Document Number", "Temp Revision Number"], dropna=False, sort=False):
        _prepass_explicit_true_pair(out, grev, target_exts)
        _prepass_strict_lvn_pairing(out, grev, is_latest_val=True)
        _prepass_strict_lvn_pairing(out, grev, is_latest_val=False)
        _handle_lone_pdfs_and_unmatched_targets(out, grev, target_exts)

    out = tidy(out)
    return out.drop(columns=["row_id", "_paired"], errors="ignore")


# ========= IO =========
def write_output_to_same_workbook(df_out: pd.DataFrame, workbook_path: str, output_sheet: str) -> None:
    if Path(workbook_path).exists():
        wb = load_workbook(workbook_path)
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        wb.save(workbook_path)
        wb.close()
        mode = "a"
    else:
        mode = "w"

    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode=mode, if_sheet_exists="replace") as w:
        df_out.to_excel(w, sheet_name=output_sheet, index=False)


def run():
    print("Starting rendition path propagation process...")
    df = pd.read_excel(WORKBOOK_PATH, sheet_name=INPUT_SHEET)
    print(f"Loaded '{INPUT_SHEET}' from {WORKBOOK_PATH} with {len(df)} rows")

    df = normalize_df(df)
    df_out = apply_spreadsheet_updates(df, TARGET_EXTS)

    counts = df_out["Action"].value_counts().to_dict()
    print("Summary:")

    for k in [
        "UPDATE RENDITION (Explicit True Pair)",
        "UPDATE RENDITION (LVN match, isLatest=True)",
        "UPDATE RENDITION (LVN match, isLatest=False)",
        "UPDATE RENDITION (LONE PDF)",
        "REVIEW",
        "REMOVE",
        "NO CHANGE",
    ]:
        if k in counts:
            print(f" {k}: {counts[k]}")

    write_output_to_same_workbook(df_out, WORKBOOK_PATH, OUTPUT_SHEET)
    finished_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"Processed sheet '{INPUT_SHEET}' → wrote '{OUTPUT_SHEET}'. Finished at {finished_at}")


if __name__ == "__main__":
    run()
