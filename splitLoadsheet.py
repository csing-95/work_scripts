# split_import_sheets.py
import sys
import os
import time
import pandas as pd
from typing import List, Tuple, Dict, Optional
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pandas.api.types import is_float_dtype

# -----------------------------
# Utilities (ports of your VBA)
# -----------------------------
def does_exist(path: str) -> bool:
    return os.path.exists(path)

def version_id_to_path(version_id: str) -> str:
    """Loose Python port of VersionID2Path. Returns '' on parse error."""
    try:
        h1 = int(version_id[20:24] + version_id[25:29], 16)
        h2 = int(version_id[29:37], 16)
        h3 = int(version_id[10:14] + version_id[15:19], 16)
        nib = int(version_id[6], 16)
        seg = (nib & 0xC)
        cont = (nib & 0x3)
        part4 = f"{int(version_id[1:6],16):X}{int(seg):X}00"
        path = os.sep + os.path.join(
            f"{h1:X}", f"{h2:X}", f"{h3:X}", part4, f"cont.{cont}{version_id[7:9]}",
        )
        return path
    except Exception:
        return ""

# -----------------------------
# Formatting config
# -----------------------------
DATE_FMT = "DD/MM/YYYY"
FLOAT_FMT = "0.0############"   # keep at least one decimal
LEGACY_DEC_COL = "Legacy Version Number"
DATE_COLUMNS = {"Content Approved On", "Original Created On", "Content Reviewed On", "Due Date"}
IGNORE_COLUMNS = {"Year", "Rev.", "Document-Type"}

# -----------------------------
# Chunk helpers
# -----------------------------
def _chunk_by_stack(df: pd.DataFrame, stack_col: str, target_size: int) -> Tuple[List[Tuple[int, int]], Dict[str, int]]:
    blocks = []
    cur_stack = None
    start = None
    last = None

    for i, sid in zip(df.index, df[stack_col]):
        sid_key = ("__NA__", i) if pd.isna(sid) else ("ID", sid)
        if cur_stack is None:
            cur_stack = sid_key; start = i; last = i
        elif sid_key == cur_stack:
            last = i
        else:
            blocks.append((cur_stack, start, last, last - start + 1))
            cur_stack = sid_key; start = i; last = i
    if cur_stack is not None:
        blocks.append((cur_stack, start, last, last - start + 1))

    chunks: List[Tuple[int, int]] = []
    chunk_start: Optional[int] = None
    chunk_end: Optional[int] = None
    running = 0

    for _, b_start, b_end, b_size in blocks:
        if running == 0:
            chunk_start = b_start; chunk_end = b_end; running = b_size
        elif running + b_size <= target_size:
            chunk_end = b_end; running += b_size
        else:
            chunks.append((chunk_start, chunk_end))
            chunk_start = b_start; chunk_end = b_end; running = b_size

    if running > 0 and chunk_start is not None and chunk_end is not None:
        chunks.append((chunk_start, chunk_end))

    max_block = max((b[3] for b in blocks), default=0)
    oversized_blocks = [b for b in blocks if b[3] > target_size]
    diagnostics = {"max_stack_size": int(max_block), "oversized_stack_count": int(len(oversized_blocks))}
    return chunks, diagnostics

# -----------------------------
# Minimal type prep + formatting
# -----------------------------
def _prepare_chunk_types(df_chunk: pd.DataFrame) -> pd.DataFrame:
    """
    Only:
      - coerce Legacy Version Number to numeric
      - coerce DATE_COLUMNS to datetime (day-first)
    Everything else is left untouched.
    """
    out = df_chunk.copy()

    if LEGACY_DEC_COL in out.columns:
        try:
            out[LEGACY_DEC_COL] = pd.to_numeric(out[LEGACY_DEC_COL], errors="coerce")
        except Exception:
            pass

    for col in (DATE_COLUMNS & set(out.columns)):
        try:
            out[col] = pd.to_datetime(out[col], dayfirst=True, errors="coerce")
        except Exception:
            pass

    return out

def _apply_excel_number_formats(ws, df_chunk: pd.DataFrame):
    """
    Only apply formats to:
      - Legacy Version Number -> FLOAT_FMT
      - DATE_COLUMNS -> DATE_FMT
    """
    cols = list(df_chunk.columns)

    # Legacy Version Number
    if LEGACY_DEC_COL in cols:
        col_idx = cols.index(LEGACY_DEC_COL) + 1
        letter = get_column_letter(col_idx)
        # if it's numeric-like or has been coerced, set format anyway
        if (LEGACY_DEC_COL in df_chunk and is_float_dtype(df_chunk[LEGACY_DEC_COL])) or True:
            for row in range(2, ws.max_row + 1):
                ws[f"{letter}{row}"].number_format = FLOAT_FMT

    # Date columns
    for col in (DATE_COLUMNS & set(cols)):
        col_idx = cols.index(col) + 1
        letter = get_column_letter(col_idx)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = DATE_FMT

# -----------------------------
# Core logic
# -----------------------------
def split_import_sheets(source_xlsx: str, rows_per_sheet: int, base_name: str, imp_code_prefix: str):
    start_time = time.time()
    src_dir = os.path.dirname(os.path.abspath(source_xlsx))

    print("=" * 60)
    print("[INIT] Split Import Sheets")
    print("=" * 60)
    print(f"[INFO] Source workbook     : {source_xlsx}")
    print(f"[INFO] Output directory     : {src_dir}")
    print(f"[INFO] Rows per sheet (goal): {rows_per_sheet}")
    print(f"[INFO] Base file name       : {base_name}")
    print(f"[INFO] Import code prefix   : {imp_code_prefix}")
    print("-" * 60)

    # Load
    print("[INFO] Loading first sheet...")
    df = pd.read_excel(source_xlsx, sheet_name=0)
    total_rows = len(df)
    print(f"[INFO] Loaded {total_rows} rows and {len(df.columns)} columns.")

    # Required columns
    if "Document Number" not in df.columns:
        raise ValueError("Column 'Document Number' not found in the source sheet.")
    if "Import Code" not in df.columns:
        df.insert(len(df.columns), "Import Code", "")

    doc_col = "Document Number"
    imp_col = "Import Code"
    stack_col = "Stack ID" if "Stack ID" in df.columns else None

    # Decide chunk plan
    if stack_col:
        print(f"[INFO] Using stack-aware splitting on column: '{stack_col}' (stacks kept intact)")
        chunks, diag = _chunk_by_stack(df, stack_col, rows_per_sheet)
        if diag["oversized_stack_count"] > 0:
            print(f"[WARN] {diag['oversized_stack_count']} stack(s) exceed target size ({rows_per_sheet}).")
        print(f"[INFO] Planned {len(chunks)} chunk(s) using stack packing.")
    else:
        print("[INFO] 'Stack ID' not found. Falling back to boundary extension by 'Document Number'.")
        chunks = []
        current = 0
        while current < total_rows:
            end = min(current + rows_per_sheet - 1, total_rows - 1)
            while end + 1 < total_rows and df.loc[end, doc_col] == df.loc[end + 1, doc_col]:
                end += 1
            chunks.append((df.index[current], df.index[end]))
            current = end + 1
        diag = {"max_stack_size": 0, "oversized_stack_count": 0}
        print(f"[INFO] Planned {len(chunks)} chunk(s) by boundary extension.")

    # Process chunks
    files: List[str] = []
    processed_rows = 0
    last_pct = -1

    for idx, (start_i, end_i) in enumerate(chunks, start=1):
        size = end_i - start_i + 1
        processed_rows += size
        pct = int((processed_rows / total_rows) * 100)
        if pct != last_pct:
            print(f"[PROGRESS] {processed_rows}/{total_rows} rows ({pct}%)")
            last_pct = pct

        print(f"\n[STEP] Chunk {idx}: rows {start_i+1} → {end_i+1} (size {size})")

        chunk = df.loc[start_i:end_i].copy()
        chunk[imp_col] = f"{imp_code_prefix}-{idx:03d}"
        print(f"[INFO] Applied Import Code: {imp_code_prefix}-{idx:03d}")

        # Only prepare Legacy Version Number + DATE_COLUMNS
        chunk_for_excel = _prepare_chunk_types(chunk)

        out_path = os.path.join(src_dir, f"{base_name}-{idx:03d}.xlsx")
        print(f"[INFO] Writing: {out_path}")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            sheet_name = "Documents"
            chunk_for_excel.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            # Only format Legacy Version Number + DATE_COLUMNS
            _apply_excel_number_formats(ws, chunk_for_excel)

            # create Excel Table over used range
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName="Documents", ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        print(f"[SUCCESS] Saved chunk {idx} → {out_path}")
        files.append(out_path)

    dur = time.time() - start_time

    # Final summary
    print("\n" + "=" * 60)
    print(" SPLIT SUMMARY")
    print("=" * 60)
    print(f"Source file            : {source_xlsx}")
    print(f"Total rows processed   : {total_rows}")
    print(f"Target rows per sheet  : {rows_per_sheet}")
    if stack_col:
        print(f"Stack-aware column     : {stack_col}")
        print(f"Max stack size         : {diag['max_stack_size']}")
        print(f"Oversized stacks       : {diag['oversized_stack_count']}")
    print(f"Workbooks created      : {len(files)}")
    print(f"Saved to directory     : {src_dir}")
    print("-" * 60)
    for i, p in enumerate(files, 1):
        print(f"[{i:02d}] {p}")
    print("-" * 60)
    print(f"Elapsed time           : {dur:.2f}s")
    print("=" * 60 + "\n")

# -----------------------------
# Interactive entrypoint
# -----------------------------
if __name__ == "__main__":
    print("=" * 60)
    print(" Split Import Sheets (interactive mode) ")
    print("=" * 60)

    # helpers to tidy pasted inputs (remove surrounding quotes)
    def _clean_input(s: str) -> str:
        return s.strip().strip('"').strip("'")

    # Source path
    src = _clean_input(input("Path to source Excel file: "))
    while not does_exist(src):
        print("[ERROR] File not found, try again.")
        src = _clean_input(input("Path to source Excel file: "))

    # Rows per sheet
    while True:
        rows_str = input("How many rows per sheet? ").strip()
        if rows_str.isdigit() and int(rows_str) > 0:
            rows = int(rows_str)
            break
        print("[ERROR] Please enter a positive integer.")

    # Base name (trim quotes if pasted)
    base = _clean_input(input("Base name for output files (e.g. MyFile): "))
    while not base:
        print("[ERROR] Base name cannot be empty.")
        base = _clean_input(input("Base name for output files (e.g. MyFile): "))

    # Import code prefix (trim quotes if pasted)
    imp = _clean_input(input("Import code prefix (e.g. IMP20250918): "))
    while not imp:
        print("[ERROR] Import code prefix cannot be empty.")
        imp = _clean_input(input("Import code prefix (e.g. IMP20250918): "))

    try:
        split_import_sheets(src, rows, base, imp)
    except Exception as e:
        print(f"[ERROR] {e}")
        sys.exit(3)
