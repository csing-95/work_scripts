#!/usr/bin/env python3
from __future__ import annotations
from pathlib import Path
from typing import List, Optional
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# ========= CONFIG =========
# WORKBOOK_PATH = r"filepath/filename.xlsx"
WORKBOOK_PATH = r"C:\Users\corri\Kinsmen Group\P25025 Plains Adept to Meridian Migration Project - Documents\6. Data Migration\Location Loadsheets\Projects\Working\000 - Plains Midstream Canada Projects Loadsheet.xlsx"
INPUT_SHEET   = "Documents"
OUTPUT_SHEET  = "Rendition_Actions"
TARGET_EXTS   = ["dwg", "doc", "docx", "xls", "xlsm", "xlsx", "gp4"]
REVIEW_TAG    = "REVIEW: MATCH UNCLEAR"
HEADER_RENAME_MAP = {
    "stack_id": "Stack ID",
    "document number": "Document Number",
    "revision number": "Revision Number",
    "legacy version number": "Legacy Version Number",
    "ext": "Ext",
    "source_path": "Source Path",
    "rendition_path": "Rendition Path",
    "islatest": "isLatest",
}

# ========= HELPERS =========
def _norm_rev(val) -> str:
    s = str(val).strip()
    if s == "" or s.lower() == "nan": return ""
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else s
    except ValueError:
        pass
    return str(int(s)) if s.isdigit() else s.upper()

def _truthy(x) -> bool:
    if pd.isna(x): return False
    if isinstance(x, bool): return x
    return str(x).strip().lower() in {"true","yes","y","1","t"}

def _is_adlib_path(s: pd.Series) -> pd.Series:
    return s.astype(str).str.contains(r"\\ADLib_", case=False, regex=True).fillna(False)

def _pick_pdf_row_id(gdf: pd.DataFrame) -> Optional[int]:
    if gdf.empty: return None
    g = gdf.copy()
    g["_is_adlib"] = _is_adlib_path(g["Source Path"]).astype(int)
    g = g.sort_values(["Legacy Version Number", "_is_adlib", "row_id"])  # take last
    return int(g.iloc[-1]["row_id"])

def _pick_target_row_id(gdf: pd.DataFrame, target_exts: List[str]) -> Optional[int]:
    if gdf.empty: return None
    pr = {e: i for i, e in enumerate([e.lower() for e in target_exts])}
    g = gdf.copy()
    g["ext_rank"] = g["Ext"].str.lower().map(pr).fillna(len(pr)).astype(int)
    g = g.sort_values(["ext_rank", "Legacy Version Number", "row_id"])  # take first
    return int(g.iloc[0]["row_id"])

def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns={c: c.strip() for c in df.columns})
    df = df.rename(columns={c: HEADER_RENAME_MAP[c.lower()] for c in df.columns if c.lower() in HEADER_RENAME_MAP})
    required = {"Stack ID","Document Number","Revision Number","Legacy Version Number","Ext","Source Path","Rendition Path"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns: {sorted(missing)}")

    df = df.copy()
    for c in ["Ext","Revision Number","Document Number","Stack ID","Source Path","Rendition Path"]:
        df[c] = df[c].astype(object).where(pd.notna(df[c]), "")
        df[c] = df[c].astype(str).str.strip()

    df["isLatest"] = df["isLatest"].apply(_truthy) if "isLatest" in df.columns else False
    df["Revision Number"] = df["Revision Number"].apply(_norm_rev)
    df["Ext"] = df["Ext"].str.lower().str.replace(".", "", regex=False)
    df["Legacy Version Number"] = pd.to_numeric(df["Legacy Version Number"], errors="coerce").fillna(-1).astype(int)
    df["row_id"] = df.index
    return df

def tidy(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["Rendition Path","Source Path","Stack ID","Document Number","Revision Number"]:
        df[col] = df[col].replace({"nan":"","NaN":""})
    return df

# Writers
def _write_update_with_path(out: pd.DataFrame, rid: int, action: str, copied_path: str):
    out.at[rid, "Action"] = action
    out.at[rid, "Reason"] = ""
    out.at[rid, "Rendition Path"] = copied_path or ""

def _write_reason_only(out: pd.DataFrame, rid: int, action: str, reason: str):
    out.at[rid, "Action"] = action
    out.at[rid, "Reason"] = reason
    out.at[rid, "Rendition Path"] = reason

# ========= CORE =========
def _prepass_islatest_pairs_per_revision(out: pd.DataFrame, target_exts: List[str]) -> None:
    """Within each (Stack, Document, Revision), if exactly two isLatest=True rows (PDF + target), pair them."""
    out["_paired"] = False if "_paired" not in out.columns else out["_paired"]
    target_exts = [e.lower().lstrip(".") for e in target_exts]

    for rkey, grev in out.groupby(["Stack ID","Document Number","Revision Number"], dropna=False, sort=False):
        latest = grev[grev["isLatest"]]
        n = len(latest)
        if n == 2:
            pdfs = latest[latest["Ext"].eq("pdf")]
            tgts = latest[latest["Ext"].isin(target_exts)]
            if len(pdfs) == 1 and len(tgts) == 1:
                pdf_id = int(pdfs.iloc[0]["row_id"])
                tgt_id = int(tgts.iloc[0]["row_id"])
                src = out.at[pdf_id, "Source Path"]
                _write_update_with_path(out, tgt_id, "UPDATE RENDITION (isLatest pair)", src)
                _write_reason_only(out, pdf_id, "REMOVE", "REMOVE: PDF consumed by isLatest target")
                out.at[pdf_id, "_paired"] = True
                out.at[tgt_id, "_paired"] = True
            else:
                for rid in latest["row_id"]:
                    _write_reason_only(out, int(rid), "REVIEW", f"{REVIEW_TAG}: isLatest pair not PDF+target")
        elif n > 2:
            for rid in latest["row_id"]:
                _write_reason_only(out, int(rid), "REVIEW", f"{REVIEW_TAG}: >2 isLatest in same revision")

def _pair_within_revision_by_legacy(out: pd.DataFrame, grev: pd.DataFrame, target_exts: List[str]) -> bool:
    """
    Try to pair PDF + target that share the SAME Legacy Version Number within this revision.
    Returns True if at least one pair was made.
    """
    made_pair = False
    target_exts = [e.lower().lstrip(".") for e in target_exts]
    g_unpaired = grev[~grev["_paired"].fillna(False)] if "_paired" in grev.columns else grev

    # Work LVNs in descending order to prefer newer legacy numbers
    for lvn, g_lvn in g_unpaired.groupby("Legacy Version Number", sort=False):
        if pd.isna(lvn): 
            continue
        pdfs = g_lvn[g_lvn["Ext"].eq("pdf")]
        tgts = g_lvn[g_lvn["Ext"].isin(target_exts)]
        if len(pdfs) and len(tgts):
            pdf_id = _pick_pdf_row_id(pdfs)
            tgt_id = _pick_target_row_id(tgts, target_exts)
            if pdf_id is not None and tgt_id is not None:
                src = out.at[pdf_id, "Source Path"]
                _write_update_with_path(out, tgt_id, "UPDATE RENDITION (LVN match)", src)
                _write_reason_only(out, pdf_id, "REMOVE", "REMOVE: PDF consumed by LVN-matched target")
                out.at[pdf_id, "_paired"] = True
                out.at[tgt_id, "_paired"] = True
                made_pair = True

                # extras within same LVN group → review
                for idx in pdfs.index.difference([pdf_id]):
                    _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra PDF with same LVN")
                for idx in tgts.index.difference([tgt_id]):
                    _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra target with same LVN")
    return made_pair

def apply_spreadsheet_updates(df: pd.DataFrame, target_exts: List[str]) -> pd.DataFrame:
    target_exts = [e.lower().lstrip(".") for e in target_exts]
    out = df.copy()
    out["Action"] = "NO CHANGE"
    out["Reason"] = ""
    out["_paired"] = False

    # 1) Pre-pass: isLatest pairing BY REVISION
    _prepass_islatest_pairs_per_revision(out, target_exts)

    # 2) Main pass: per revision
    for rkey, grev in out.groupby(["Stack ID","Document Number","Revision Number"], dropna=False, sort=False):
        # First try LVN pairing within this revision
        paired_any = _pair_within_revision_by_legacy(out, grev, target_exts)

        # Recompute unpaired after LVN step
        g_unpaired = out.loc[grev.index]
        g_unpaired = g_unpaired[~g_unpaired["_paired"].fillna(False)]

        if g_unpaired.empty:
            continue

        pdfs = g_unpaired[g_unpaired["Ext"].str.lower().eq("pdf")]
        tgts = g_unpaired[g_unpaired["Ext"].isin(target_exts)]

        if len(pdfs) and len(tgts):
            # Fallback: best PDF vs best target in this revision
            pdf_id = _pick_pdf_row_id(pdfs)
            tgt_id = _pick_target_row_id(tgts, target_exts)
            if pdf_id is not None and tgt_id is not None:
                src = out.at[pdf_id, "Source Path"]
                _write_update_with_path(out, tgt_id, "UPDATE RENDITION", src)
                _write_reason_only(out, pdf_id, "REMOVE", "REMOVE: PDF consumed by target")
                out.at[pdf_id, "_paired"] = True
                out.at[tgt_id, "_paired"] = True
            else:
                for rid in g_unpaired["row_id"]:
                    _write_reason_only(out, int(rid), "REVIEW", f"{REVIEW_TAG}: picker failure")

            # extras in same revision → review
            for idx in pdfs.index.difference([pdf_id] if 'pdf_id' in locals() and pdf_id is not None else []):
                _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra PDF in revision")
            for idx in tgts.index.difference([tgt_id] if 'tgt_id' in locals() and tgt_id is not None else []):
                _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra target in revision")

        elif len(pdfs) and not len(tgts):
            pdf_id = _pick_pdf_row_id(pdfs)
            if pdf_id is not None:
                _write_update_with_path(out, pdf_id, "UPDATE RENDITION (LONE PDF)", out.at[pdf_id, "Source Path"])
                out.at[pdf_id, "_paired"] = True
            for idx in pdfs.index.difference([pdf_id] if pdf_id is not None else []):
                _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra PDF without target")

        elif not len(pdfs) and len(tgts):
            tgt_id = _pick_target_row_id(tgts, target_exts)
            if tgt_id is not None:
                _write_reason_only(out, int(tgt_id), "REVIEW", f"{REVIEW_TAG}: no PDF for this revision")
            for idx in tgts.index.difference([tgt_id] if tgt_id is not None else []):
                _write_reason_only(out, int(idx), "REVIEW", f"{REVIEW_TAG}: extra target without PDF")

        # else: neither → leave as NO CHANGE

    out = tidy(out)
    return out.drop(columns=["row_id", "_paired"], errors="ignore")

# ========= IO =========
def write_output_to_same_workbook(df_out: pd.DataFrame, workbook_path: str, output_sheet: str) -> None:
    if Path(workbook_path).exists():
        wb = load_workbook(workbook_path)
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        wb.save(workbook_path); wb.close()
        mode = "a"
    else:
        mode = "w"
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode=mode, if_sheet_exists="replace") as w:
        df_out.to_excel(w, sheet_name=output_sheet, index=False)

def run():
    print("Starting rendition path propagation process...")
    df = pd.read_excel(WORKBOOK_PATH, sheet_name=INPUT_SHEET)
    print(f"Loaded '{INPUT_SHEET}' from {WORKBOOK_PATH} with {len(df)} rows")
    df = normalize_df(df)
    df_out = apply_spreadsheet_updates(df, TARGET_EXTS)

    counts = df_out["Action"].value_counts().to_dict()
    print("Summary:")
    for k in ["UPDATE RENDITION (isLatest pair)","UPDATE RENDITION (LVN match)","UPDATE RENDITION","UPDATE RENDITION (LONE PDF)","REVIEW","REMOVE","NO CHANGE"]:
        if k in counts:
            print(f"  {k}: {counts[k]}")

    write_output_to_same_workbook(df_out, WORKBOOK_PATH, OUTPUT_SHEET)
    finished_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"Processed sheet '{INPUT_SHEET}' → wrote '{OUTPUT_SHEET}'. Finished at {finished_at}")

if __name__ == "__main__":
    run()

