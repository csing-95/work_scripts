import os
import threading
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

try:
    import sv_ttk
    HAS_SV_TTK = True
except ImportError:
    HAS_SV_TTK = False


def norm_header(v):
    if v is None:
        return ""
    return str(v).strip()


def build_sheet_list(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    names = wb.sheetnames
    wb.close()
    return names


def read_headers_fast(xlsx_path, sheet_name):
    """
    Fast header read using iter_rows (no random ws.cell calls).
    Returns: (headers_in_order, header_to_col_index_1based)
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in: {xlsx_path}")

    ws = wb[sheet_name]

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        header_row = row
        break

    if header_row is None:
        wb.close()
        return [], {}

    headers = []
    hmap = {}
    for idx, cell in enumerate(header_row, start=1):
        h = norm_header(cell.value)
        if h:
            headers.append(h)
            # if duplicates exist, keep the first
            if h not in hmap:
                hmap[h] = idx

    wb.close()
    return headers, hmap


def merge_excels_fast(
    files,
    sheet_name,
    output_path,
    add_source_file=True,
    dedupe=False,
    status_cb=None,
):
    if not files:
        raise ValueError("No files selected.")
    if not sheet_name:
        raise ValueError("No sheet selected.")
    if not output_path:
        raise ValueError("No output path set.")

    # 1) Union headers across files (by name)
    all_headers = []
    file_header_maps = []  # (file, header->source_col)

    for i, f in enumerate(files, start=1):
        if status_cb:
            status_cb(f"Scanning headers ({i}/{len(files)}): {os.path.basename(f)}")
        hdrs, hmap = read_headers_fast(f, sheet_name)
        file_header_maps.append((f, hmap))
        for h in hdrs:
            if h not in all_headers:
                all_headers.append(h)

    if add_source_file and "Source File" not in all_headers:
        all_headers.append("Source File")

    # output column index for each header
    out_col_for_header = {h: j for j, h in enumerate(all_headers, start=1)}

    # 2) Create output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = f"Merged_{sheet_name}"[:31]

    # write header row
    for h, j in out_col_for_header.items():
        out_ws.cell(row=1, column=j).value = h

    out_row = 2
    seen = set()

    # 3) Merge each file (fast row iteration)
    for i, (f, hmap) in enumerate(file_header_maps, start=1):
        if status_cb:
            status_cb(f"Merging ({i}/{len(file_header_maps)}): {os.path.basename(f)}")

        wb = load_workbook(f, read_only=True, data_only=False)
        ws = wb[sheet_name]

        # Build a mapping from source column index -> output column index
        # Only for headers that exist in this source file.
        src_to_out = {}
        for h, src_col in hmap.items():
            out_col = out_col_for_header.get(h)
            if out_col is not None:
                src_to_out[src_col] = out_col

        # For dedupe signature, we use the union headers (excluding Source File)
        sig_headers = [h for h in all_headers if h != "Source File"]

        # Iterate rows starting row 2
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            # quick skip: if row is totally empty (common)
            if not any(c.value is not None and c.value != "" for c in row_cells):
                continue

            if dedupe:
                sig = []
                for h in sig_headers:
                    src_col = hmap.get(h)
                    if src_col is None:
                        sig.append(None)
                    else:
                        # row_cells is 0-based
                        sig.append(row_cells[src_col - 1].value)
                sig = tuple(sig)
                if sig in seen:
                    continue
                seen.add(sig)

            # copy values + number formats
            for src_col, out_col in src_to_out.items():
                src_cell = row_cells[src_col - 1]
                dst_cell = out_ws.cell(row=out_row, column=out_col)
                dst_cell.value = src_cell.value
                dst_cell.number_format = src_cell.number_format

            if add_source_file:
                out_ws.cell(row=out_row, column=out_col_for_header["Source File"]).value = os.path.basename(f)

            out_row += 1

        wb.close()

    # Optional: set a sensible default width (doesn't alter values)
    for col in range(1, len(all_headers) + 1):
        out_ws.column_dimensions[get_column_letter(col)].width = 18

    out_wb.save(output_path)
    if status_cb:
        status_cb(f"Done ✅ Saved: {output_path}")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Merge Excel Sheets (Preserve Values)")
        self.geometry("900x580")

        self.files = []

        self.sheet_var = tk.StringVar(value="")
        self.add_source_var = tk.BooleanVar(value=True)
        self.dedupe_var = tk.BooleanVar(value=False)
        self.output_var = tk.StringVar(value="")

        self._build_ui()

        if HAS_SV_TTK:
            try:
                sv_ttk.set_theme("dark")
            except Exception:
                pass

    def _build_ui(self):
        main = ttk.Frame(self, padding=12)
        main.pack(fill="both", expand=True)

        # Files
        files_frame = ttk.LabelFrame(main, text="Input files", padding=10)
        files_frame.pack(fill="x")

        btns = ttk.Frame(files_frame)
        btns.pack(fill="x")

        ttk.Button(btns, text="Add Excel files...", command=self.add_files).pack(side="left")
        ttk.Button(btns, text="Remove selected", command=self.remove_selected).pack(side="left", padx=8)
        ttk.Button(btns, text="Clear", command=self.clear_files).pack(side="left")

        self.listbox = tk.Listbox(files_frame, height=6, selectmode=tk.EXTENDED)
        self.listbox.pack(fill="x", pady=8)

        # Sheet selection
        sheet_frame = ttk.LabelFrame(main, text="Sheet/tab to merge", padding=10)
        sheet_frame.pack(fill="x", pady=10)

        row = ttk.Frame(sheet_frame)
        row.pack(fill="x")

        ttk.Label(row, text="Sheet:").pack(side="left")
        self.sheet_combo = ttk.Combobox(row, textvariable=self.sheet_var, state="readonly", width=40, values=[])
        self.sheet_combo.pack(side="left", padx=8)

        ttk.Button(row, text="Refresh from first file", command=self.refresh_sheets).pack(side="left")

        # Options
        opt_frame = ttk.LabelFrame(main, text="Options", padding=10)
        opt_frame.pack(fill="x")

        ttk.Checkbutton(opt_frame, text="Add 'Source File' column", variable=self.add_source_var).pack(anchor="w")
        ttk.Checkbutton(opt_frame, text="Remove duplicates (exact match across columns)", variable=self.dedupe_var).pack(anchor="w")

        # Output
        out_frame = ttk.LabelFrame(main, text="Output", padding=10)
        out_frame.pack(fill="x", pady=10)

        out_row = ttk.Frame(out_frame)
        out_row.pack(fill="x")

        ttk.Entry(out_row, textvariable=self.output_var).pack(side="left", fill="x", expand=True)
        ttk.Button(out_row, text="Choose save location...", command=self.choose_output).pack(side="left", padx=8)

        # Run
        run_frame = ttk.Frame(main)
        run_frame.pack(fill="x", pady=6)

        self.run_btn = ttk.Button(run_frame, text="Merge", command=self.run_merge)
        self.run_btn.pack(side="left")

        self.status = ttk.Label(run_frame, text="")
        self.status.pack(side="left", padx=12)

        # Log
        log_frame = ttk.LabelFrame(main, text="Log", padding=10)
        log_frame.pack(fill="both", expand=True)

        self.log = tk.Text(log_frame, height=10, wrap="word")
        self.log.pack(fill="both", expand=True)

    def log_line(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log.insert("end", f"[{ts}] {msg}\n")
        self.log.see("end")
        self.status.config(text=msg)
        self.update_idletasks()

    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select Excel files",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not paths:
            return
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.listbox.insert("end", p)
        self.refresh_sheets()

    def remove_selected(self):
        sel = list(self.listbox.curselection())
        if not sel:
            return
        for idx in reversed(sel):
            path = self.listbox.get(idx)
            self.listbox.delete(idx)
            if path in self.files:
                self.files.remove(path)
        self.refresh_sheets()

    def clear_files(self):
        self.files = []
        self.listbox.delete(0, "end")
        self.sheet_combo["values"] = []
        self.sheet_var.set("")

    def refresh_sheets(self):
        if not self.files:
            self.sheet_combo["values"] = []
            self.sheet_var.set("")
            return
        first = self.files[0]
        try:
            names = build_sheet_list(first)
            self.sheet_combo["values"] = names
            if "Documents" in names:
                self.sheet_var.set("Documents")
            else:
                self.sheet_var.set(names[0] if names else "")
            self.log_line(f"Loaded sheets from: {os.path.basename(first)}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not read sheet names:\n{e}")

    def choose_output(self):
        path = filedialog.asksaveasfilename(
            title="Save merged file as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            self.output_var.set(path)

    def run_merge(self):
        if not self.files:
            messagebox.showwarning("Missing files", "Please add at least one Excel file.")
            return
        sheet = self.sheet_var.get().strip()
        if not sheet:
            messagebox.showwarning("Missing sheet", "Please select a sheet/tab to merge.")
            return
        out_path = self.output_var.get().strip()
        if not out_path:
            messagebox.showwarning("Missing output", "Please choose an output file path.")
            return

        self.run_btn.config(state="disabled")

        def worker():
            try:
                self.log_line("Starting merge...")
                merge_excels_fast(
                    files=self.files,
                    sheet_name=sheet,
                    output_path=out_path,
                    add_source_file=self.add_source_var.get(),
                    dedupe=self.dedupe_var.get(),
                    status_cb=self.log_line
                )
                self.after(0, lambda: messagebox.showinfo("Done", f"Merged file saved:\n{out_path}"))
            except Exception as e:
                tb = traceback.format_exc()
                self.log_line("ERROR")
                self.log_line(str(e))
                self.log_line(tb)
                self.after(0, lambda: messagebox.showerror("Merge failed", str(e)))
            finally:
                self.after(0, lambda: self.run_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    app = App()
    if HAS_SV_TTK:
        try:
            sv_ttk.set_theme("dark")
        except Exception:
            pass
    app.mainloop()